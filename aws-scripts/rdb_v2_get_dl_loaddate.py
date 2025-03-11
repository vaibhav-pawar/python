import pandas as pd
import openpyxl
import boto3
import time
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from tqdm import tqdm
from concurrent.futures import ThreadPoolExecutor

athena_client = boto3.client('athena')
glue_client = boto3.client('glue')
dyno_client = boto3.client('dynamodb')

def glue_collect_database_information():
    print("\n\U000027A1 Collecting Database Names from Glue.")
    paginator = glue_client.get_paginator('get_databases')
    databases = []

    for page in paginator.paginate():
        for database in page['DatabaseList']:
            databases.append(database['Name'])

    return databases

databases = glue_collect_database_information()

rdbv2_databases_collection = [name for name in databases if name.startswith("<database-name>__")]
print("\033[32mV2 Databases:\033[0m", rdbv2_databases_collection)

print("\n\U000027A1 Collecting Table Names from Database Names.")
def get_all_glue_tables(database_name):
    paginator = glue_client.get_paginator('get_tables')
    all_tables = []

    for page in paginator.paginate(DatabaseName=database_name):
        all_tables.extend(page['TableList'])

    return all_tables

with ThreadPoolExecutor() as executor:
    result = list(executor.map(get_all_glue_tables, rdbv2_databases_collection))
print("\033[32mTable Names Collected successfully from Glue.\033[0m")

all_collected_tables_of_v2 = [table for sublist in result for table in sublist]

print("\n\U000027A1 Restructuring Database & Table Names.")
v2_tables_raw_dict = {}
for item in all_collected_tables_of_v2:
    database_name = item['DatabaseName']
    table_name = item['Name']

    if database_name not in v2_tables_raw_dict:
        v2_tables_raw_dict[database_name] = []

    v2_tables_raw_dict[database_name].append(table_name)

print("\033[32mDatabase & Table Names restructured successfully.\033[0m")

print("\n\U000027A1 Checking CDC Table to exclude tables from v2 list.")
dyno_response = dyno_client.get_item(
    TableName='<dynamo-table-name>',
    Key={
            'id': {
                'S': '<dynamo-record-name>'
            }
        }
)

cdc_raw_list = []
cdc_raw_list.append(dyno_response['Item']['table']['M']['rdb']['M'])

def convert_to_lowercase(cdc_raw_list):
    if isinstance(cdc_raw_list, str):
        return cdc_raw_list.lower()
    if isinstance(cdc_raw_list, list):
        return [convert_to_lowercase(item) for item in cdc_raw_list]
    if isinstance(cdc_raw_list, dict):
        return {key.lower(): convert_to_lowercase(value) for key, value in cdc_raw_list.items()}
    return cdc_raw_list

cdc_lowercase_data = convert_to_lowercase(cdc_raw_list)

print("\033[32mRemoving CDC table from 'v2_tables_raw_dict' List.\033[0m")
for comparing_tables_cdc_to_rdb_v2 in cdc_lowercase_data:
    for database, tables in comparing_tables_cdc_to_rdb_v2.items():
        convert_to_rdb_v2 = "<v2-database-name>__" + database
        for table in tables['ss']:
            if table in v2_tables_raw_dict[f'{convert_to_rdb_v2}']:
                v2_tables_raw_dict[f'{convert_to_rdb_v2}'].remove(f"{table}")

print("\033[32mCDC Tables are removed from v2 List.\033[0m")

print("\n\U000027A1 Exporting v2 Database & Table Names in excel File.")
v2_database_table_dict = []
for database, tables in v2_tables_raw_dict.items():
    for table in tables:
        custom_dict = {"v2_Database": database, "v2_Table": table}
        v2_database_table_dict.append(custom_dict)

df = pd.DataFrame(v2_database_table_dict)
excel_file_path = 'RDB_v2_Table_excluded_cdc.xlsx'
writer = pd.ExcelWriter(excel_file_path, engine='openpyxl')
df.to_excel(writer, index=False, sheet_name='RDB_v2_dl_loaddate')
workbook = writer.book
worksheet = writer.sheets['RDB_v2_dl_loaddate']
for cell in worksheet['1']:
    cell.fill = PatternFill(start_color='4CBB17', end_color='4CBB17', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
writer.save()
print(f"\033[32mFile saved as {excel_file_path}\033[0m")

print("\n\U000027A1 Going to sleep mode for 5 sec", end='\n')
time.sleep(5)

def run_query():
    file = "RDB_v2_Table_excluded_cdc.xlsx"
    print(f"\n\U000027A1 Collecting resources from '{file}' for data collection")
    df = pd.read_excel(file, sheet_name='RDB_v2_dl_loaddate', header=0, engine='openpyxl')
    workbook = openpyxl.load_workbook(file)

    print(f"\033[32mCollecting dl_loaddate of v2 tables\033[0m")
    for ind, row in tqdm(df.iterrows(), total=len(df), desc="Processing"):
        try:
            query_v2 = f"""
                    SELECT MAX(dl_loaddate) as v2_dl_loaddate FROM "{row["v2_Database"]}"."{row["v2_Table"]}";
                """
            query_process_v2 = athena_client.start_query_execution(QueryString=query_v2)
            query_execution_id_v2 = query_process_v2['QueryExecutionId']

            if query_execution_id_v2:
                while True:
                    query_execution_v2 = athena_client.get_query_execution(QueryExecutionId=query_execution_id_v2)
                    status_v2 = query_execution_v2.get('QueryExecution', {}).get('Status', {}).get('State')
                    StateChangeReason_v2 = query_execution_v2.get('QueryExecution', {}).get('Status', {}).get(
                        'StateChangeReason')

                    if status_v2 == "SUCCEEDED":
                        break
                    elif status_v2 in ["QUEUED", "RUNNING"]:
                        time.sleep(2)
                    else:
                        print(f"Athena query status change reason: {StateChangeReason_v2}")
                        break

                response_result_v2 = athena_client.get_query_results(QueryExecutionId=query_execution_id_v2)
                v2_dl_loaddate = response_result_v2['ResultSet']['Rows'][1]['Data'][0]['VarCharValue']

            df.at[ind, 'v2_dl_loaddate'] = v2_dl_loaddate
        except Exception as e:
            print(f"Error occured on Database - {row['v2_Database']} Table - {row['v2_Table']} , {str(e)}")

    print(f"\033[32mv2 table 'dl_loaddate' collected successfully.\033[0m")

    print("\n\U000027A1 Exporting table dl_loaddate in excel file.")
    output_excel_file = 'RDB_v2_dl_loaddate.xlsx'
    writer = pd.ExcelWriter(output_excel_file, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='RDB_v2_dl_loaddate')
    workbook = writer.book
    worksheet = writer.sheets['RDB_v2_dl_loaddate']
    for cell in worksheet['1']:
        cell.fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    writer.save()
    print(f"\033[32mFile saved as {output_excel_file}\033[0m")

run_query()