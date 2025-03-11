import boto3
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from tqdm import tqdm
import botocore
import json
import time
import sys
from concurrent.futures import ThreadPoolExecutor

client = boto3.client('glue')
glue_client = boto3.client('glue')
athena_client = boto3.client('athena')
dyno_client = boto3.client('dynamodb')

print("---------------------------------------------------")
print("PROCESS 1 : Resource Collection is started.")
print("---------------------------------------------------")

def glue_collect_database_information():
    print("\n\U000027A1 Collecting Database Names from Glue.")
    paginator = glue_client.get_paginator('get_databases')
    databases = []

    for page in paginator.paginate():
        for database in page['DatabaseList']:
            databases.append(database['Name'])

    return databases

databases = glue_collect_database_information()

rdb_databases_collection = [name for name in databases if name.startswith("<v1-database>_")]
rdbv2_databases_collection = [name for name in databases if name.startswith("<v2-database>__")]

print("\033[32mV1 Databases:\033[0m", rdb_databases_collection)
print("\033[32mV2 Databases:\033[0m", rdbv2_databases_collection)

if len(rdb_databases_collection) == len(rdbv2_databases_collection):
    v1_databases_with_prefix = [f'<v1-database>__{db.split("_", 1)[1]}' for db in rdb_databases_collection]
    v2_databases_without_prefix = [db.replace('v2_', '') for db in rdbv2_databases_collection]

    missing_in_v1 = [db for db in v2_databases_without_prefix if db not in rdb_databases_collection]
    missing_in_v2 = [db for db in v1_databases_with_prefix if db not in rdbv2_databases_collection]

    if not missing_in_v1 and not missing_in_v2:
        print("\033[32mPassed: Database is present is both sides [V1 & V2]\033[0m")
    else:
        if missing_in_v1:
            print("\033[31mMissing values in V1 databases:\033[0m", missing_in_v1)
        if missing_in_v2:
            print("\033[31mMissing values in V2 databases:\033[0m", missing_in_v2)
else:
    sys.exit("\033[31mFailed: Database count does not match\033[0m")
print("\033[32mDatabase Names Collected successfully from Glue.\033[0m")

print("\n\U000027A1 Collecting Table Names from Database Names.")
def get_all_glue_tables(database_name):
    paginator = glue_client.get_paginator('get_tables')
    all_tables = []

    for page in paginator.paginate(DatabaseName=database_name):
        all_tables.extend(page['TableList'])

    return all_tables

with ThreadPoolExecutor() as executor:
    result1 = list(executor.map(get_all_glue_tables, rdb_databases_collection))
    result2 = list(executor.map(get_all_glue_tables, rdbv2_databases_collection))
print("\033[32mTable Names Collected successfully from Glue.\033[0m")

all_collected_tables_of_v1 = [table for sublist in result1 for table in sublist]
all_collected_tables_of_v2 = [table for sublist in result2 for table in sublist]

print("\n\U000027A1 Restructuring Database & Table Names.")
v1_tables_raw_dict = {}
v2_tables_raw_dict = {}

for item in all_collected_tables_of_v1:
    database_name = item['DatabaseName']
    table_name = item['Name']

    if database_name not in v1_tables_raw_dict:
        v1_tables_raw_dict[database_name] = []

    v1_tables_raw_dict[database_name].append(table_name)

for item in all_collected_tables_of_v2:
    database_name = item['DatabaseName']
    table_name = item['Name']

    if database_name not in v2_tables_raw_dict:
        v2_tables_raw_dict[database_name] = []

    v2_tables_raw_dict[database_name].append(table_name)

print("\033[32mDatabase & Table Names restructured successfully.\033[0m")

print("\n\U000027A1 Checking CDC Table to exclude tables from v1 & v2 list.")
dyno_response = dyno_client.get_item(
    TableName='<dynamo-table-name>',
    Key={
            'id': {
                'S': '<tables-list-key-name'
            }
        }
)

cdc_raw_list = []
cdc_raw_list.append(dyno_response['Item']['table']['M']['rdb']['M'])

def convert_to_lowercase(cdc_raw_list):
    if isinstance(cdc_raw_list, str):
        return cdc_raw_list.lower()
    if isinstance(cdc_raw_list, list):
        return [convert_to_lowercase(item) for item in cdc_raw_list]
    if isinstance(cdc_raw_list, dict):
        return {key.lower(): convert_to_lowercase(value) for key, value in cdc_raw_list.items()}
    return cdc_raw_list

cdc_lowercase_data = convert_to_lowercase(cdc_raw_list)

print("\033[32mRemoving CDC table from 'v1_tables_raw_dict' List.\033[0m")
for comparing_tables_cdc_to_rdb_v1 in cdc_lowercase_data:
    for database, tables in comparing_tables_cdc_to_rdb_v1.items():
        convert_to_rdb_v1 = "<v1-database-starting-string>_" + database
        for table in tables['ss']:
            if table in v1_tables_raw_dict[f'{convert_to_rdb_v1}']:
                v1_tables_raw_dict[f'{convert_to_rdb_v1}'].remove(f"{table}")

print("\033[32mRemoving CDC table from 'v2_tables_raw_dict' List.\033[0m")
for comparing_tables_cdc_to_rdb_v2 in cdc_lowercase_data:
    for database, tables in comparing_tables_cdc_to_rdb_v2.items():
        convert_to_rdb_v2 = "<v2-database-starting-string>__" + database
        for table in tables['ss']:
            if table in v2_tables_raw_dict[f'{convert_to_rdb_v2}']:
                v2_tables_raw_dict[f'{convert_to_rdb_v2}'].remove(f"{table}")

print("\033[32mCDC Tables are removed from v1 & v2 List.\033[0m")

# print(json.dumps(v1_tables_raw_dict, indent=4))
# print(json.dumps(v2_tables_raw_dict, indent=4))

print("\n\U000027A1 Combining v1 & v2 Database + table in one dictionary.")
v1_v2_combined = []

for database, tables in v2_tables_raw_dict.items():
    split_v2 = database.split('v2_')
    v2_name_modify_to_v1 = split_v2[0] + split_v2[1]
    for table in tables:
        if table in v1_tables_raw_dict.get(v2_name_modify_to_v1, []):
            custom_dict = {"v1_Database": v2_name_modify_to_v1, "v1_Table": table, "v2_Database": database, "v2_Table": table}
            v1_v2_combined.append(custom_dict)
        else:
            custom_dict = {"v1_Database": v2_name_modify_to_v1, "v1_Table": "Table_Not_Found", "v2_Database": database, "v2_Table": table}
            v1_v2_combined.append(custom_dict)

for database, tables in v1_tables_raw_dict.items():
    split_v2 = database.split('_')
    v1_name_modify_to_v2 = split_v2[0] + "v2__" + split_v2[1]
    for table in tables:
        if table in v2_tables_raw_dict.get(v1_name_modify_to_v2, []):
            custom_dict = {"v1_Database": database, "v1_Table": table, "v2_Database": v1_name_modify_to_v2, "v2_Table": table}
            v1_v2_combined.append(custom_dict)
        else:
            custom_dict = {"v1_Database": database, "v1_Table": table, "v2_Database": v1_name_modify_to_v2, "v2_Table": "Table_Not_Found"}
            v1_v2_combined.append(custom_dict)

unique_tuples = set(tuple(d.items()) for d in v1_v2_combined)
unique_dicts = [dict(t) for t in unique_tuples]
sorted_dicts = sorted(unique_dicts, key=lambda x: x['v1_Database'])
print("\033[32mv1 & v2 tables Combined Successfully.\033[0m")

print("\n\U000027A1 Exporting combined Database & Table Names in excel File.")
df = pd.DataFrame(sorted_dicts)
excel_file_path = 'RDB_v1_v2_Table_excluded_cdc.xlsx'
writer = pd.ExcelWriter(excel_file_path, engine='openpyxl')
df.to_excel(writer, index=False, sheet_name='RDB_v1_v2')
workbook = writer.book
worksheet = writer.sheets['RDB_v1_v2']
for cell in worksheet['1']:
    cell.fill = PatternFill(start_color='4CBB17', end_color='4CBB17', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
writer.save()
print(f"\033[32mFile saved as {excel_file_path}\033[0m")

print("\n\U000027A1 Going to sleep mode for 5 sec", end='\n')
time.sleep(5)

print("--------------------------------------------------------")
print("PROCESS 2 : Resource data collection is started.")
print("--------------------------------------------------------")

def run_query():
    file = "RDB_v1_v2_Table_excluded_cdc.xlsx"
    print(f"\n\U000027A1 Collecting resources from '{file}' for data collection")
    df = pd.read_excel(file, sheet_name='RDB_v1_v2', header=0, engine='openpyxl')
    workbook = openpyxl.load_workbook(file)

    print(f"\033[32mCollecting count of v1 & v2 tables\033[0m")
    for ind, row in tqdm(df.iterrows(), total=len(df), desc="Processing"):
        if row["v1_Table"] == "Table_Not_Found":
            df.at[ind, 'v1_count'] = None
        else:
            try:
                query_v1 = f"""
                        SELECT count(*) as v1_count FROM "{row["v1_Database"]}"."{row["v1_Table"]}";
                    """
                query_process_v1 = athena_client.start_query_execution(QueryString=query_v1)
                query_execution_id_v1 = query_process_v1['QueryExecutionId']

                if query_execution_id_v1:
                    while True:
                        query_execution_v1 = athena_client.get_query_execution(QueryExecutionId=query_execution_id_v1)
                        status_v1 = query_execution_v1.get('QueryExecution', {}).get('Status', {}).get('State')
                        StateChangeReason_v1 = query_execution_v1.get('QueryExecution', {}).get('Status', {}).get(
                            'StateChangeReason')

                        if status_v1 == "SUCCEEDED":
                            break
                        elif status_v1 in ["QUEUED", "RUNNING"]:
                            time.sleep(2)
                        else:
                            print(f"Error for v1: {StateChangeReason_v1}")
                            break

                    response_result_v1 = athena_client.get_query_results(QueryExecutionId=query_execution_id_v1)
                    v1_count = int(response_result_v1['ResultSet']['Rows'][1]['Data'][0]['VarCharValue'])

                df.at[ind, 'v1_count'] = v1_count
            except Exception as e:
                print(f"Error for v1: {str(e)}")

        if row["v2_Table"] == "Table_Not_Found":
            df.at[ind, 'v2_count'] = None
        else:
            try:
                query_v2 = f"""
                        SELECT count(*) as v2_count FROM "{row["v2_Database"]}"."{row["v2_Table"]}";
                    """
                query_process_v2 = athena_client.start_query_execution(QueryString=query_v2)
                query_execution_id_v2 = query_process_v2['QueryExecutionId']

                if query_execution_id_v2:
                    while True:
                        query_execution_v2 = athena_client.get_query_execution(QueryExecutionId=query_execution_id_v2)
                        status_v2 = query_execution_v2.get('QueryExecution', {}).get('Status', {}).get('State')
                        StateChangeReason_v2 = query_execution_v2.get('QueryExecution', {}).get('Status', {}).get(
                            'StateChangeReason')

                        if status_v2 == "SUCCEEDED":
                            break
                        elif status_v2 in ["QUEUED", "RUNNING"]:
                            time.sleep(2)
                        else:
                            print(f"Error for v2: {StateChangeReason_v2}")
                            break

                    response_result_v2 = athena_client.get_query_results(QueryExecutionId=query_execution_id_v2)
                    v2_count = int(response_result_v2['ResultSet']['Rows'][1]['Data'][0]['VarCharValue'])

                df.at[ind, 'v2_count'] = v2_count
            except Exception as e:
                print(f"Error for v2: {str(e)}")

    print(f"\033[32mv1 & v2 table count collected successfully.\033[0m")

    print(f"\n\U000027A1 Comparing count of v1 & v2")
    df["Status"] = df["v1_count"] == df["v2_count"]
    df["Status"] = df["Status"].astype(str)
    print(f"\033[32mCount comparison completed.\033[0m")

    print("\n\U000027A1 Exporting table count & status data in excel file.")
    output_excel_file = 'RDB_v1_v2_Checkup.xlsx'
    writer = pd.ExcelWriter(output_excel_file, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='RDB_v1_v2')
    workbook = writer.book
    worksheet = writer.sheets['RDB_v1_v2']
    for cell in worksheet['1']:
        cell.fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    writer.save()
    print(f"\033[32mFile saved as {output_excel_file}\033[0m")

run_query()